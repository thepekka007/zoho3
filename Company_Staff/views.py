from django.db import models
from django.shortcuts import render,redirect
from Register_Login.models import *
from Company_Staff.models import *
from Register_Login.views import logout
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.contrib import messages
from django.core.mail import send_mail
from django.core.mail import EmailMessage
from xhtml2pdf import pisa
from django.template.loader import get_template
from bs4 import BeautifulSoup
import io
from openpyxl import Workbook
from openpyxl import load_workbook
from django.http import HttpResponse,HttpResponseRedirect
from django.conf import settings
from io import BytesIO
# Create your views here.
from datetime import date
from django.db.models import Max

from django.db.models import Q
# -------------------------------Company section--------------------------------

# company dashboard
def company_dashboard(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        context = {
            'details': dash_details,
            'allmodules': allmodules
        }
        return render(request, 'company/company_dash.html', context)
    else:
        return redirect('/')


# company profile
def company_profile(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        context = {
            'details': dash_details,
            'allmodules': allmodules
        }
        return render(request, 'company/company_profile.html', context)
    else:
        return redirect('/')

# company profile
def company_staff_request(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        staff_request=StaffDetails.objects.filter(company=dash_details.id, company_approval=0).order_by('-id')
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'requests':staff_request,
        }
        return render(request, 'company/staff_request.html', context)
    else:
        return redirect('/')

# company staff accept or reject
def staff_request_accept(request,pk):
    staff=StaffDetails.objects.get(id=pk)
    staff.company_approval=1
    staff.save()
    return redirect('company_staff_request')

def staff_request_reject(request,pk):
    staff=StaffDetails.objects.get(id=pk)
    login_details=LoginDetails.objects.get(id=staff.company.id)
    login_details.delete()
    staff.delete()
    return redirect('company_staff_request')


# All company staff view, cancel staff approval
def company_all_staff(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        all_staffs=StaffDetails.objects.filter(company=dash_details.id, company_approval=1).order_by('-id')
        print(all_staffs)
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'staffs':all_staffs,
        }
        return render(request, 'company/all_staff_view.html', context)
    else:
        return redirect('/')

def staff_approval_cancel(request,pk):
    staff=StaffDetails.objects.get(id=pk)
    staff.company_approval=2
    staff.save()
    return redirect('company_all_staff')









# -------------------------------Staff section--------------------------------

# staff dashboard
def staff_dashboard(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        context={
            'details':dash_details,
            'allmodules': allmodules,
        }
        return render(request,'staff/staff_dash.html',context)
    else:
        return redirect('/')


# staff profile
def staff_profile(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
           
        else:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        context={
            'details':dash_details,
            'allmodules': allmodules,
        }
        return render(request,'staff/staff_profile.html',context)
    else:
        return redirect('/')











# -------------------------------Zoho Modules section--------------------------------
    



    #--------------------------------------------------- TINTO VIEW ITEMS START-------------------------------------------

# items llist
    
def items_list(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        login_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=login_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                item=Items.objects.filter(company=dash_details.company)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'item':item,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/items/items_list.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            item=Items.objects.filter(company=dash_details)
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            context = {
                    'details': dash_details,
                    'item': item,
                    'allmodules': allmodules,
            }
        return render(request,'zohomodules/items/items_list.html',context)

   
   
# create Items

def new_items(request):                                                                #new by tinto mt
    if 'login_id' in request.session:
        login_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=login_id)
    if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                item=Items.objects.filter(company=dash_details.company)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                units = Unit.objects.filter(company=dash_details.company)
                accounts=Chart_of_Accounts.objects.filter(company=dash_details.company)
                context = {
                     'details': dash_details,
                    'units': units,
                    'allmodules': allmodules,
                    'accounts':accounts
                }
                return render(request,'zohomodules/items/newitem.html',context)
    if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            item=Items.objects.filter(company=dash_details)
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            units = Unit.objects.filter(company=dash_details)
            accounts=Chart_of_Accounts.objects.filter(company=dash_details)
            context = {
                    'details': dash_details,
                    'units': units,
                    'allmodules': allmodules,
                    'accounts':accounts
            }
    
            return render(request, 'zohomodules/items/newitem.html',context)
# create Items
def create_item(request):                                                                #new by tinto mt
    
    login_id = request.session['login_id']
    if 'login_id' not in request.session:
        return redirect('/')
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
        company_id = request.session['login_id']
        
        if request.method=='POST':
            a=Items()
            b=Item_Transaction_History()
            c = CompanyDetails.objects.get(login_details=company_id)
            b.company=c
            b.logindetails=log_user
            a.login_details=log_user
            a.company=c
            a.item_type = request.POST.get("type",None)
            a.item_name = request.POST.get("name",None)
            unit_id = request.POST.get("unit")
            uid=Unit.objects.get(id=unit_id)
            # unit_instance = get_object_or_404(Unit, id=unit_id)
            a.unit = uid
            a.hsn_code = request.POST.get("hsn",None)
            a.tax_reference = request.POST.get("radio",None)
            a.intrastate_tax = request.POST.get("intra",None)
            a.interstate_tax= request.POST.get("inter",None)
            a.selling_price = request.POST.get("sel_price",None)
            a.sales_account = request.POST.get("sel_acc",None)
            a.sales_description = request.POST.get("sel_desc",None)
            a.purchase_price = request.POST.get("cost_price",None)
            a.purchase_account = request.POST.get("cost_acc",None)
            a.purchase_description = request.POST.get("pur_desc",None)
            a.minimum_stock_to_maintain = request.POST.get("minimum_stock",None)
            a.activation_tag = 'Active'
            a.type = 'Opening Stock'
            a.inventory_account = request.POST.get("invacc",None)
            a.opening_stock = request.POST.get("openstock",None)
            a.opening_stock_per_unit = request.POST.get("rate",None)
            
            a.save()    
            t=Items.objects.get(id=a.id)
            b.items=t
            b.save()
            return redirect('items_list')
    elif log_user.user_type == 'Staff':
        staff_id = request.session['login_id']
        if request.method=='POST':
            a=Items()
            b=Item_Transaction_History()
            staff = LoginDetails.objects.get(id=staff_id)
            sf = StaffDetails.objects.get(login_details=staff)
            c=sf.company
            b.Date=date.today()
            b.company=c
            b.logindetails=log_user
            a.login_details=log_user
            a.company=c
            a.item_type = request.POST.get("type",None)
            a.item_name = request.POST.get("name",None)
            unit_id = request.POST.get("unit")
            unit_instance = get_object_or_404(Unit, id=unit_id)
            a.unit = unit_instance
            a.hsn_code = request.POST.get("hsn",None)
            a.tax_reference = request.POST.get("radio",None)
            a.intrastate_tax = request.POST.get("intra",None)
            a.interstate_tax= request.POST.get("inter",None)
            a.selling_price = request.POST.get("sel_price",None)
            a.sales_account = request.POST.get("sel_acc",None)
            a.sales_description = request.POST.get("sel_desc",None)
            a.purchase_price = request.POST.get("cost_price",None)
            a.purchase_account = request.POST.get("cost_acc",None)
            a.purchase_description = request.POST.get("pur_desc",None)
            a.minimum_stock_to_maintain = request.POST.get("minimum_stock",None)
            a.activation_tag = request.POST.get("status",None)
            a.inventory_account = request.POST.get("invacc",None)
            a.opening_stock = request.POST.get("openstock",None)
        
        

        
            a.save()    
            t=Items.objects.get(id=a.id)
            b.items=t
            b.save()
            return redirect('items_list')
    return redirect('items_list')

# create unit
def add_unit(request):                                                                #new by tinto mt (item)
    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)

    if log_user.user_type == 'Company':
        if request.method == 'POST':
            c = CompanyDetails.objects.get(login_details=login_id)
            unit_name = request.POST['units']
            
            if Unit.objects.filter(unit_name=unit_name, company=c).exists():
                return JsonResponse({"message": "error"})
            else:
                unit = Unit(unit_name=unit_name, company=c)  
                unit.save()  
                return JsonResponse({"message": "success"})

    elif log_user.user_type == 'Staff':
        if request.method == 'POST':
            staff = LoginDetails.objects.get(id=login_id)
            sf = StaffDetails.objects.get(login_details=staff)
            c = sf.company
            unit_name = request.POST['units']
            
            if Unit.objects.filter(unit_name=unit_name, company=c).exists():
                return JsonResponse({"message": "error"})
            else:
                unit = Unit(unit_name=unit_name, company=c)  
                unit.save()  
                return JsonResponse({"message": "success"})

    return JsonResponse({"message": "success"})
# create unit


    
def unit_dropdown(request):                                                                 #new by tinto mt (item)
    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_user)
            options = {}
            option_objects = Unit.objects.filter(company=dash_details)
            for option in option_objects:
                unit_name=option.unit_name
            options[option.id] = [unit_name,f"{unit_name}"]
            return JsonResponse(options)
      

    elif log_user.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_user)
            options = {}
            option_objects = Unit.objects.filter(company=dash_details.company)
            for option in option_objects:
                unit_name=option.unit_name
            options[option.id] = [unit_name,f"{unit_name}"]
            return JsonResponse(options)
             



def add_account(request):                                                                #new by tinto mt (item)
    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
        company_id = request.session['login_id']
        if request.method == 'POST':
            a=Chart_of_Accounts()
            b=Chart_of_Accounts_History()
            c = CompanyDetails.objects.get(login_details=company_id)
            b.company=c
            b.logindetails=log_user
            b.action="Created"
            b.Date=date.today()
            a.login_details=log_user
            a.company=c
          
        
            a.account_type = request.POST.get("account_type",None)
            a.account_name = request.POST.get("account_name",None)
            a.account_code = request.POST.get("account_code",None)
            a.description = request.POST.get("description",None)
    
            a.Create_status="active"
            ac_name=request.POST.get("account_name",None)
            if Chart_of_Accounts.objects.filter(account_name=ac_name, company=c).exists():
                return JsonResponse({"message": "error"})
            else:
          
                a.save()
                t=Chart_of_Accounts.objects.get(id=a.id)
                b.chart_of_accounts=t
                b.save()
                acc_id = a.id  
                acc_name=a.account_name
                response_data = {
                "message": "success",
                "acc_id":acc_id,
                "acc_name":acc_name,
        
                         }

                return JsonResponse(response_data)
        

    elif log_user.user_type == 'Staff':
        staff_id = request.session['login_id']
        if request.method=='POST':
            a=Chart_of_Accounts()
            b=Chart_of_Accounts_History()
            staff = LoginDetails.objects.get(id=staff_id)
            sf = StaffDetails.objects.get(login_details=staff)
            a=sf.company
            b.Date=date.today()
            b.company=c
            b.logindetails=log_user
            a.login_details=log_user
            a.company=c
          
        
            a.account_type = request.POST.get("account_type",None)
            a.account_name = request.POST.get("account_name",None)
            a.account_code = request.POST.get("account_code",None)
            a.description = request.POST.get("description",None)
    
            a.Create_status="active"
            ac_name=request.POST.get("account_name",None)
            if Chart_of_Accounts.objects.filter(account_name=ac_name, company=c).exists():
                return JsonResponse({"message": "error"})
            else:
          
                a.save()
                t=Chart_of_Accounts.objects.get(id=a.id)
                b.chart_of_accounts=t
                b.save()
                acc_id = a.id  
                acc_name=a.account_name
                response_data = {
                "message": "success",
                "acc_id":acc_id,
                "acc_name":acc_name,
        
                         }

                return JsonResponse(response_data)
        
      
        
    return redirect('newitems')

def account_dropdown(request):                                                                #new by tinto mt (item)
    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_user)
            options = {}
            option_objects = Chart_of_Accounts.objects.filter(Q(company=dash_details) & (Q(account_type='Expense') | Q(account_type='Other Expense') | Q(account_type='Cost Of Goods Sold')))
            for option in option_objects:
                account_name=option.account_name
                account_type=option.account_type
                options[option.id] = [account_name,f"{account_name}"]
            return JsonResponse(options)
    elif log_user.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_user)
            options = {}
       
            option_objects = Chart_of_Accounts.objects.filter(Q(company=dash_details.company) & (Q(account_type='Expense') | Q(account_type='Other Expense') | Q(account_type='Cost Of Goods Sold')))
            for option in option_objects:
                account_name=option.account_name
                options[option.id] = [account_name,f"{account_name}"]
            return JsonResponse(options)


    
    
def itemsoverview(request,pk):                                                                 #new by tinto mt

    if 'login_id' in request.session:
        login_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=login_id)
    if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                item=Items.objects.filter(company=dash_details.company)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
              
                items=Items.objects.filter(company=dash_details.company)
                selitem=Items.objects.get(id=pk)
                est_comments=Items_comments.objects.filter(Items=pk)
                stock_value=selitem.opening_stock*selitem.purchase_price  
                latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
                filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)
                context = {
                     'details': dash_details,
                
                    'allmodules': allmodules,
                    'items':items,
                    'selitem':selitem,
                    'stock_value':stock_value,
                    'latest_item_id':filtered_data,
                    'est_comments':est_comments
                }
                return render(request, 'zohomodules/items/itemsoverview.html',context)
    if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
       
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            items=Items.objects.filter(company=dash_details)
            selitem=Items.objects.get(id=pk)
            est_comments=Items_comments.objects.filter(Items=pk)
            stock_value=selitem.opening_stock*selitem.purchase_price  
            latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
            filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)
            context = {
                    'details': dash_details,
                   
                    'allmodules': allmodules,
                    'items':items,
                    'selitem':selitem,
                    'stock_value':stock_value,
                    'latest_item_id':filtered_data,
                    'est_comments':est_comments
            }
    
            return render(request, 'zohomodules/items/itemsoverview.html',context)


    return render(request, 'zohomodules/items/itemsoverview.html')


def item_status_edit(request, pv):                                                                #new by tinto mt
    
    selitem = Items.objects.get(id=pv)

    if selitem.activation_tag == 'Active':
        selitem.activation_tag = 'inactive'
        selitem.save()
    elif selitem.activation_tag != 'Active':
        selitem.activation_tag = 'Active'
        selitem.save()

    selitem.save()

    return redirect('itemsoverview',pv)


def shareItemToEmail(request,pt):                                                                #new by tinto mt
    if request.user: 
        try:
            if request.method == 'POST':
                emails_string = request.POST['email_ids']
                # Split the string by commas and remove any leading or trailing whitespace
                emails_list = [email.strip() for email in emails_string.split(',')]
                email_message = request.POST['email_message']
                print(emails_list)
                print('1')
           
           
                item = Items.objects.get(id=pt)
                context = {
                
                    'selitem':item,
                }
                print('2')
                template_path = 'zohomodules/items/itememailpdf.html'
                print('3')
                template = get_template(template_path)
                print('4')
                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)#, link_callback=fetch_resources)
                pdf = result.getvalue()
                print('5')
                filename = f'Item Transactions.pdf'
                subject = f"Transactipns"
                email = EmailMessage(subject, f"Hi,\nPlease find the attached Item transactions. \n{email_message}\n\n--\nRegards,\n{item.item_name}\n{item.item_type}", from_email=settings.EMAIL_HOST_USER,to=emails_list)
                email.attach(filename,pdf,"application/pdf")
                email.send(fail_silently=False)
                msg = messages.success(request, 'Details has been shared via email successfully..!')
                return redirect(itemsoverview,pt)
        except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect(itemsoverview,pt)   
        
def deleteitem(request,pl):                                                                #new by tinto mt
    items=Items.objects.filter(id=pl)
    items.delete()
    
    return redirect(items_list)




def add_item_comment(request,pc):                                                                #new by tinto mt

    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
        company_id = request.session['login_id']
        if request.method=="POST":
                    
                    com=Items_comments()
                    c = CompanyDetails.objects.get(login_details=company_id)
            
                    comment_comments=request.POST['comment']
                    com.company=c
                    com.logindetails=log_user
                    com.comments=comment_comments
                    item=Items.objects.get(id=pc)
                    com.Items=item
                    
                    com.save()
                    return redirect('itemsoverview',pc)

    elif log_user.user_type == 'Staff':
        staff_id = request.session['login_id']
        if request.method=='POST':
            com=Items_comments()
            staff = LoginDetails.objects.get(id=staff_id)
            sf = StaffDetails.objects.get(login_details=staff)
            c=sf.company
            
            comment_comments=request.POST['comment']
            com.company=c
            com.logindetails=log_user
            com.comments=comment_comments
            item=Items.objects.get(id=pc)
            com.Items=item
                    
            com.save()
            return redirect('itemsoverview',pc)
    return redirect('itemsoverview',pc)
        




         
def downloadItemSampleImportFile(request):                                                                  #new by tinto mt
    estimate_table_data = [['No.','ITEM TYPE','ITEM NAME','HSN','TAX REFERENCE','INTRASTATE TAX','INTERSTATE TAX','SELLING PRICE','SALES ACCOUNT','SALES DESCRIPTION','PURCHASE PRICE','PURCHASE ACCOUNT','PURCHASE DESCRIPTION','MINIMUM STOCK TO MAINTAIN','ACTIVATION TAG','OPENING STOCK','CURRENT STOCK','OPENING STOCK PER UNIT']]      
    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = 'Sheet1'
    

    # Populate the sheets with data
    for row in estimate_table_data:
        sheet1.append(row)  
    
    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=expense_sample_file.xlsx'
     # Save the workbook to the response
    wb.save(response)
    return response





def import_item(request):                                                                #new by tinto mt
    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)

    if log_user.user_type == 'Company':
        company_id = request.session['login_id']

        if request.method == 'POST' and 'excel_file' in request.FILES:
            company = CompanyDetails.objects.get(login_details=company_id)
            excel_file = request.FILES.get('excel_file')
            wb = load_workbook(excel_file)

            try:
                ws = wb["Sheet1"]
                header_row = ws[1]
                column_names = [cell.value for cell in header_row]
                print("Column Names:", column_names)
            except KeyError:
                print('Sheet not found')
                messages.error(request, '`Sheet1` not found in the Excel file. Please check.')
                return redirect('expensepage')

            expected_columns = ['No.', 'ITEM TYPE', 'ITEM NAME', 'HSN', 'TAX REFERENCE', 'INTRASTATE TAX', 'INTERSTATE TAX',
                                'SELLING PRICE', 'SALES ACCOUNT', 'SALES DESCRIPTION', 'PURCHASE PRICE',
                                'PURCHASE ACCOUNT', 'PURCHASE DESCRIPTION', 'MINIMUM STOCK TO MAINTAIN', 'ACTIVATION TAG',
                                'OPENING STOCK', 'CURRENT STOCK', 'OPENING STOCK PER UNIT']

            if column_names != expected_columns:
                print('Invalid sheet columns or order')
                messages.error(request, 'Sheet column names or order is not in the required format. Please check.')
                return redirect("comapny_items")

            for row in ws.iter_rows(min_row=2, values_only=True):
                _, item_type, item_name, hsn, tax_reference, intrastate_tax, interstate_tax, selling_price, sales_account, \
                sales_description, purchase_price, purchase_account, purchase_description, min_stock, activation_tag, \
                opening_stock, current_stock, opening_stock_per_unit = row

                # Fetching the 'Unit' instance with id=1 (you may adjust this based on your 'Unit' model)
                unit_instance = Unit.objects.get(pk=1)

                # Creating an instance of the 'Items' model and saving it
                item = Items(
                    login_details=log_user,
                    company=company,
                    unit=unit_instance,  # Use the fetched 'Unit' instance
                    item_type=item_type,
                    item_name=item_name,
                    hsn_code=hsn,
                    tax_reference=tax_reference,
                    intrastate_tax=intrastate_tax,
                    interstate_tax=interstate_tax,
                    selling_price=selling_price,
                    sales_account=sales_account,
                    sales_description=sales_description,
                    purchase_price=purchase_price,
                    purchase_account=purchase_account,
                    purchase_description=purchase_description,
                    minimum_stock_to_maintain=min_stock,
                    activation_tag=activation_tag,
                    inventory_account="Inventory Account",
                    opening_stock=opening_stock,
                    opening_stock_per_unit=opening_stock_per_unit
                )
                item.save()

            messages.success(request, 'Data imported successfully!')
            return redirect("items_list")
        else:
            messages.error(request, 'Invalid request. Please check the file and try again.')
            return redirect("items_list")
    else:
        messages.error(request, 'Invalid user type. Please check your user type.')
        return redirect("items_list")


def item_view_sort_by_name(request, pk):                                                                #new by tinto mt
    # Retrieve all items and convert them to a list of dictionaries
    items = list(Items.objects.all().values())

    # Sort the items by the 'item_name' field
    sorted_items = sorted(items, key=lambda r: r['item_name'])

    # Get the selected item by ID
    selitem = Items.objects.get(id=pk)

    # Fetch related comments for the selected item
    est_comments = Items_comments.objects.filter(Items=pk)

    # Calculate stock value for the selected item
    stock_value = selitem.opening_stock * selitem.purchase_price

    # Find the latest date for the item transaction history
    latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']

    # Filter transaction history for the latest date and the selected item
    filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)

    # Render the template with the sorted items and other relevant data
    return render(request, 'zohomodules/items/itemsoverview.html', {'items': sorted_items, 'selitem': selitem, 'stock_value': stock_value, 'latest_item_id': filtered_data, 'est_comments': est_comments})

def item_view_sort_by_hsn(request, pk):                                                                #new by tinto mt
    # Retrieve all items and convert them to a list of dictionaries
    items = list(Items.objects.all().values())

    # Sort the items by the 'item_name' field
    sorted_items = sorted(items, key=lambda r: r['hsn_code'])

    # Get the selected item by ID
    selitem = Items.objects.get(id=pk)

    # Fetch related comments for the selected item
    est_comments = Items_comments.objects.filter(Items=pk)

    # Calculate stock value for the selected item
    stock_value = selitem.opening_stock * selitem.purchase_price

    # Find the latest date for the item transaction history
    latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']

    # Filter transaction history for the latest date and the selected item
    filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)

    # Render the template with the sorted items and other relevant data
    return render(request, 'zohomodules/items/itemsoverview.html', {'items': sorted_items, 'selitem': selitem, 'stock_value': stock_value, 'latest_item_id': filtered_data, 'est_comments': est_comments})

def filter_item_view_Active(request,pk):                                                                #new by tinto mt
    items=Items.objects.filter(activation_tag='Active')  
    selitem=Items.objects.get(id=pk)
    est_comments=Items_comments.objects.filter(Items=pk)
    stock_value=selitem.opening_stock*selitem.purchase_price  
    latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
    filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)

    return render(request, 'zohomodules/items/itemsoverview.html',{'items':items,'selitem':selitem,'stock_value':stock_value,'latest_item_id':filtered_data,'est_comments':est_comments})

def filter_item_view_inActive(request,pk):                                                                #new by tinto mt
    items=Items.objects.filter(activation_tag='inactive')  
    selitem=Items.objects.get(id=pk)
    est_comments=Items_comments.objects.filter(Items=pk)
    stock_value=selitem.opening_stock*selitem.purchase_price  
    latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
    filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)

    return render(request, 'zohomodules/items/itemsoverview.html',{'items':items,'selitem':selitem,'stock_value':stock_value,'latest_item_id':filtered_data,'est_comments':est_comments})

    
    #--------------------------------------------------- TINTO VIEW ITEMS END-------------------------------------------


        #--------------------------------------------------- TINTO VIEW CHART OF ACCOUNTS START-------------------------------------------
def addchartofaccounts(request):                                                                #new by tinto mt
        if 'login_id' in request.session:
            login_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=login_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                item=Items.objects.filter(company=dash_details.company)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                units = Unit.objects.filter(company=dash_details.company)
                accounts=Chart_of_Accounts.objects.filter(company=dash_details.company)
                context = {
                     'details': dash_details,
        
                    'allmodules': allmodules,
         
                }
                return render(request,'zohomodules/chartofaccounts/addchartofaccounts.html',context)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            item=Items.objects.filter(company=dash_details)
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            units = Unit.objects.filter(company=dash_details)
            accounts=Chart_of_Accounts.objects.filter(company=dash_details)
            context = {
                    'details': dash_details,
          
                    'allmodules': allmodules,
           
            }
    
            return render(request,'zohomodules/chartofaccounts/addchartofaccounts.html',context)


def chartofaccounts(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            acc=Chart_of_Accounts.objects.filter(company=dash_details)
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
  
def chartofaccountsActive(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company,status="active")
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            acc=Chart_of_Accounts.objects.filter(company=dash_details)
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)

def chartofaccountsInactive(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company,status="inactive")
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            acc=Chart_of_Accounts.objects.filter(company=dash_details,status="inactive")
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
            
        
def create_account(request):                                                                #new by tinto mt
    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
        company_id = request.session['login_id']
        if request.method=='POST':
            a=Chart_of_Accounts()
            b=Chart_of_Accounts_History()
            account=Chart_of_Accounts.objects.all()
            c = CompanyDetails.objects.get(login_details=company_id)
            b.company=c
            b.logindetails=log_user
            b.action="Created"
            b.Date=date.today()
            a.login_details=log_user
            a.company=c
          
        
            a.account_type = request.POST.get("account_type",None)
            a.account_name = request.POST.get("account_name",None)
            a.account_code = request.POST.get("account_code",None)
            a.account_number = request.POST.get("account_number2",None)
            a.account_description = request.POST.get("description",None)
            if a.account_type=="Other Current Assets":

                a.sub_account = request.POST.get("sub_account",None)
                a.parent_account = request.POST.get("parent_account",None)
               

            if a.account_type=="Cash":
             
                a.sub_account = request.POST.get("sub_account22",None)
                a.parent_account = request.POST.get("parent_account22",None)
               

            if a.account_type=="Fixed Assets":
            
                a.sub_account = request.POST.get("sub_account33",None)
                a.parent_account = request.POST.get("parent_account33",None)
               
            
            if a.account_type=="Stock":
               
                a.sub_account = request.POST.get("sub_account44",None)
                a.parent_account = request.POST.get("parent_account44",None)
             
            
            if a.account_type=="Other Current Liability":
             
                a.sub_account = request.POST.get("sub_account55",None)
                a.parent_account = request.POST.get("parent_account55",None)
               
            if a.account_type=="Long Term Liability":
            
                a.sub_account = request.POST.get("sub_account66",None)
                a.parent_account = request.POST.get("parent_account66",None)
              
            
            if a.account_type=="Other Liability":
              
                a.sub_account = request.POST.get("sub_account77",None)
                a.parent_account = request.POST.get("parent_account77",None)
              
            if a.account_type=="Equity":
            
                a.sub_account = request.POST.get("sub_account88",None)
                a.parent_account = request.POST.get("parent_account88",None)
            
            
            if a.account_type=="Income":
             
                a.sub_account = request.POST.get("sub_account99",None)
                a.parent_account = request.POST.get("parent_account99",None)
              
            
            if a.account_type=="Expense":
             
                a.sub_account = request.POST.get("sub_account100",None)
                a.parent_account = request.POST.get("parent_account100",None)
              
            if a.account_type=="Cost Of Goods Sold":
              
                a.sub_account = request.POST.get("sub_account111",None)
                a.parent_account = request.POST.get("parent_account111",None)
             
            if a.account_type=="Other Expense":
             
                a.sub_account = request.POST.get("sub_account222",None)
                a.parent_account = request.POST.get("parent_account222",None)
               
            account_type=request.POST.get("account_type",None)
            if account_type == 'Other Assets':
                a.description = 'Track special assets like goodwill and other intangible assets'
            if account_type == 'Other Current Assets':
                a.description = 'Any short term asset that can be converted into cash or cash equivalents easily Prepaid expenses Stocks and Mutual Funds'
            if account_type == 'Cash':
                a.description = 'To keep track of cash and other cash equivalents like petty cash, undeposited funds, etc., use an organized accounting system  financial software'
            if account_type == 'Bank':
                a.description = 'To keep track of bank accounts like Savings, Checking, and Money Market accounts.'
            if account_type == 'Fixed Asset':
                a.description = 'Any long-term investment or asset that cannot be easily converted into cash includes: Land and Buildings, Plant, Machinery, and Equipment, Computers, Furniture.'
            if account_type == 'Stock':
                a.description = 'To keep track of your inventory assets.'
            if account_type == 'Payment Clearing':
                a.description = 'To keep track of funds moving in and out via payment processors like Stripe, PayPal, etc.'
            if account_type == 'Other Liability':
                a.description = 'Obligation of an entity arising from past transactions or events which would require repayment.Tax to be paid Loan to be Repaid Accounts Payableetc.'
            if account_type == 'Other Current Liability':
                a.description = 'Any short term liability like: Customer Deposits Tax Payable'
            if account_type == 'Credit Card':
                a.description = 'Create a trail of all your credit card transactions by creating a credit card account.'
            if account_type == 'Long Term Liability':
                a.description = 'Liabilities that mature after a minimum period of one year like: Notes Payable Debentures Long Term Loans '
            if account_type == 'Overseas Tax Payable':
                a.description = 'Track your taxes in this account if your business sells digital services to foreign customers.'
            if account_type == 'Equity':
                a.description = 'Owners or stakeholders interest on the assets of the business after deducting all the liabilities.'
            if account_type == 'Income':
                a.description = 'Income or Revenue earned from normal business activities like sale of goods and services to customers.'
            if account_type == 'Other Income':
                a.description = 'Income or revenue earned from activities not directly related to your business like : Interest Earned Dividend Earned'
            if account_type == 'Expense':
                a.description = 'Reflects expenses incurred for running normal business operations, such as : Advertisements and Marketing Business Travel Expenses License Fees Utility Expenses'
            if account_type == 'Cost Of Goods Sold':
                a.description = 'This indicates the direct costs attributable to the production of the goods sold by a company such as: Material and Labor costs Cost of obtaining raw materials'
            if account_type == 'Other Expense':
                a.description = 'Track miscellaneous expenses incurred for activities other than primary business operations or create additional accounts to track default expenses like insurance or contribution towards charity.'
       

            
    
            a.Create_status="active"
            ac_name=request.POST.get("account_name",None)
            if Chart_of_Accounts.objects.filter(account_name=ac_name,company=c).exists():
                error='yes'
                messages.error(request,'Account with same name exsits !!!')
                return redirect('addchartofaccounts')
            else:
                a.save()
                t=Chart_of_Accounts.objects.get(id=a.id)
                b.chart_of_accounts=t
                b.save()
                return redirect('chartofaccounts')
    elif log_user.user_type == 'Staff':
        staff_id = request.session['login_id']
        if request.method=='POST':
            a=Chart_of_Accounts()
            b=Chart_of_Accounts_History()
            staff = LoginDetails.objects.get(id=staff_id)
            sf = StaffDetails.objects.get(login_details=staff)
            c=sf.company
            b.Date=date.today()
            b.company=c
            b.logindetails=log_user
            a.login_details=log_user
            a.company=c
          
        
            a.account_type = request.POST.get("account_type",None)
            a.account_name = request.POST.get("account_name",None)
            a.account_code = request.POST.get("account_code",None)
            a.description = request.POST.get("description",None)
            if a.account_type=="Other Current Assets":

                a.sub_account = request.POST.get("sub_account",None)
                a.parent_account = request.POST.get("parent_account",None)
               

            if a.account_type=="Cash":
             
                a.sub_account = request.POST.get("sub_account22",None)
                a.parent_account = request.POST.get("parent_account22",None)
               

            if a.account_type=="Fixed Assets":
            
                a.sub_account = request.POST.get("sub_account33",None)
                a.parent_account = request.POST.get("parent_account33",None)
               
            
            if a.account_type=="Stock":
               
                a.sub_account = request.POST.get("sub_account44",None)
                a.parent_account = request.POST.get("parent_account44",None)
             
            
            if a.account_type=="Other Current Liability":
             
                a.sub_account = request.POST.get("sub_account55",None)
                a.parent_account = request.POST.get("parent_account55",None)
               
            if a.account_type=="Long Term Liability":
            
                a.sub_account = request.POST.get("sub_account66",None)
                a.parent_account = request.POST.get("parent_account66",None)
              
            
            if a.account_type=="Other Liability":
              
                a.sub_account = request.POST.get("sub_account77",None)
                a.parent_account = request.POST.get("parent_account77",None)
              
            if a.account_type=="Equity":
            
                a.sub_account = request.POST.get("sub_account88",None)
                a.parent_account = request.POST.get("parent_account88",None)
            
            
            if a.account_type=="Income":
             
                a.sub_account = request.POST.get("sub_account99",None)
                a.parent_account = request.POST.get("parent_account99",None)
              
            
            if a.account_type=="Expense":
             
                a.sub_account = request.POST.get("sub_account100",None)
                a.parent_account = request.POST.get("parent_account100",None)
              
            if a.account_type=="Cost Of Goods Sold":
              
                a.sub_account = request.POST.get("sub_account111",None)
                a.parent_account = request.POST.get("parent_account111",None)
             
            if a.account_type=="Other Expense":
             
                a.sub_account = request.POST.get("sub_account222",None)
                a.parent_account = request.POST.get("parent_account222",None)
               
            account_type=request.POST.get("account_type",None)
            if account_type == 'Other Assets':
                a.description = 'Track special assets like goodwill and other intangible assets'
            if account_type == 'Other Current Assets':
                a.description = 'Any short term asset that can be converted into cash or cash equivalents easily Prepaid expenses Stocks and Mutual Funds'
            if account_type == 'Cash':
                a.description = 'To keep track of cash and other cash equivalents like petty cash, undeposited funds, etc., use an organized accounting system  financial software'
            if account_type == 'Bank':
                a.description = 'To keep track of bank accounts like Savings, Checking, and Money Market accounts.'
            if account_type == 'Fixed Asset':
                a.description = 'Any long-term investment or asset that cannot be easily converted into cash includes: Land and Buildings, Plant, Machinery, and Equipment, Computers, Furniture.'
            if account_type == 'Stock':
                a.description = 'To keep track of your inventory assets.'
            if account_type == 'Payment Clearing':
                a.description = 'To keep track of funds moving in and out via payment processors like Stripe, PayPal, etc.'
            if account_type == 'Other Liability':
                a.description = 'Obligation of an entity arising from past transactions or events which would require repayment.Tax to be paid Loan to be Repaid Accounts Payableetc.'
            if account_type == 'Other Current Liability':
                a.description = 'Any short term liability like: Customer Deposits Tax Payable'
            if account_type == 'Credit Card':
                a.description = 'Create a trail of all your credit card transactions by creating a credit card account.'
            if account_type == 'Long Term Liability':
                a.description = 'Liabilities that mature after a minimum period of one year like: Notes Payable Debentures Long Term Loans '
            if account_type == 'Overseas Tax Payable':
                a.description = 'Track your taxes in this account if your business sells digital services to foreign customers.'
            if account_type == 'Equity':
                a.description = 'Owners or stakeholders interest on the assets of the business after deducting all the liabilities.'
            if account_type == 'Income':
                a.description = 'Income or Revenue earned from normal business activities like sale of goods and services to customers.'
            if account_type == 'Other Income':
                a.description = 'Income or revenue earned from activities not directly related to your business like : Interest Earned Dividend Earned'
            if account_type == 'Expense':
                a.description = 'Reflects expenses incurred for running normal business operations, such as : Advertisements and Marketing Business Travel Expenses License Fees Utility Expenses'
            if account_type == 'Cost Of Goods Sold':
                a.description = 'This indicates the direct costs attributable to the production of the goods sold by a company such as: Material and Labor costs Cost of obtaining raw materials'
            if account_type == 'Other Expense':
                a.description = 'Track miscellaneous expenses incurred for activities other than primary business operations or create additional accounts to track default expenses like insurance or contribution towards charity.'
       
    
            a.Create_status="active"
            ac_name=request.POST.get("account_name",None)
            if Chart_of_Accounts.objects.filter(account_name=ac_name,company=c).exists():
                error='yes'
                messages.error(request,'Account with same name exsits')
                return redirect('addchartofaccounts')
            else:
                a.save()
                t=Chart_of_Accounts.objects.get(id=a.id)
                b.chart_of_accounts=t
                b.save()
                return redirect('chartofaccounts')

    return redirect('addchartofaccounts')

def chartofaccountsoverview(request,pk):                                                                #new by tinto mt
       if 'login_id' in request.session:
        login_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=login_id)
        if log_details.user_type == 'Staff':
                    dash_details = StaffDetails.objects.get(login_details=log_details)

                    allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                
                    acc=Chart_of_Accounts.objects.filter(company=dash_details.company)  
                    selacc=Chart_of_Accounts.objects.get(id=pk)  
                    est_comments=chart_of_accounts_comments.objects.filter(chart_of_accounts=pk)
                    latest_date = Chart_of_Accounts_History.objects.filter(chart_of_accounts_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
                    filtered_data = Chart_of_Accounts_History.objects.filter(Date=latest_date, chart_of_accounts_id=pk)
                    context = {
                        'details': dash_details,
                    
                        'allmodules': allmodules,
                        'acc':acc,
                        'selacc':selacc,
                        'latest_item_id':filtered_data,
                        'est_comments':est_comments,
                    }
                    return render(request, 'zohomodules/chartofaccounts/chartofaccountsoverview.html',context)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
       
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            acc=Chart_of_Accounts.objects.filter(company=dash_details)  
            selacc=Chart_of_Accounts.objects.get(id=pk)  
            est_comments=chart_of_accounts_comments.objects.filter(chart_of_accounts=pk)
            latest_date = Chart_of_Accounts_History.objects.filter(chart_of_accounts_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
            filtered_data = Chart_of_Accounts_History.objects.filter(Date=latest_date, chart_of_accounts_id=pk)
            context = {
                        'details': dash_details,
                    
                        'allmodules': allmodules,
                        'acc':acc,
                        'selacc':selacc,
                        'latest_item_id':filtered_data,
                        'est_comments':est_comments,
                    }
    
            return render(request, 'zohomodules/chartofaccounts/chartofaccountsoverview.html',context)


   
        
    # acc=Chart_of_Accounts.objects.all()  
    # selacc=Chart_of_Accounts.objects.get(id=pk)  
    # est_comments=chart_of_accounts_comments.objects.filter(chart_of_accounts=pk)
    # latest_date = Chart_of_Accounts_History.objects.filter(chart_of_accounts_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
    # filtered_data = Chart_of_Accounts_History.objects.filter(Date=latest_date, chart_of_accounts_id=pk)
    # return render(request, 'zohomodules/chartofaccounts/chartofaccountsoverview.html',{'acc':acc,'selacc':selacc,'latest_item_id':filtered_data,'est_comments':est_comments})


from django.shortcuts import render, redirect

def editchartofaccounts(request, pr):                                                                #new by tinto mt
    # Retrieve the chart of accounts entry
    

    # Check if 'company_id' is in the session
    if 'login_id' in request.session:
        login_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    
    # Retrieve the chart of accounts entry
    acc = get_object_or_404(Chart_of_Accounts, id=pr)

    # Check if 'company_id' is in the session

    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
     
        dash_details = CompanyDetails.objects.get(login_details=log_user)
       
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
   
        context = {
                    'acc': acc,
              
                    'details': dash_details,
                   
                    'allmodules': allmodules,
            }
       
    
        
        

        if request.method == 'POST':
        
            b=Chart_of_Accounts_History()
       
            b.company=dash_details
            b.logindetails=log_user
            b.action="Edited"
            b.Date=date.today()
            acc.login_details=log_user
            acc.company=dash_details
            # Update the chart of accounts entry with the form data
            acc.account_type = request.POST['account_type']
            acc.account_name = request.POST['account_name']
            acc.account_code = request.POST['account_code']
            acc.description = request.POST['description']
            
            # Save the changes
            acc.save()
            t=Chart_of_Accounts.objects.get(id=acc.id)
            b.chart_of_accounts=t
            b.save()

            # Redirect to another page after successful update
            return redirect('chartofaccountsoverview', pr)
        return render(request, 'zohomodules/chartofaccounts/editchartofaccounts.html', context)
    if log_user.user_type == 'Staff':
        dash_details = StaffDetails.objects.get(login_details=log_user)
                
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        
   
        context = {
                    'acc': acc,
              
                    'details': dash_details,
                   
                    'allmodules': allmodules,
            }
        if request.method=='POST':
         
            b=Chart_of_Accounts_History()
         
            c=dash_details.company
            b.company=c
            b.logindetails=log_user
            b.action="Edited"
            b.Date=date.today()
            acc.login_details=log_user
            acc.company=c
            # Update the chart of accounts entry with the form data
            acc.account_type = request.POST['account_type']
            acc.account_name = request.POST['account_name']
            acc.account_code = request.POST['account_code']
            acc.description = request.POST['description']
            
            # Save the changes
            acc.save()
            t=Chart_of_Accounts.objects.get(id=acc.id)
            b.chart_of_accounts=t
            b.save()

            # Redirect to another page after successful update
            return redirect('chartofaccountsoverview', pr)
        return render(request, 'zohomodules/chartofaccounts/editchartofaccounts.html', context)

def deleteaccount(request,pl):                                                                #new by tinto mt
    acc=Chart_of_Accounts.objects.filter(id=pl)
    acc.delete()
    
    return redirect(chartofaccounts)


def acc_status_edit(request, pv):                                                                #new by tinto mt
    
    selacc = Chart_of_Accounts.objects.get(id=pv)

    if selacc.status == 'Active':
        selacc.status = 'inactive'
        selacc.save()
    elif selacc.status != 'Active':
        selacc.status = 'Active'
        selacc.save()

    selacc.save()

    return redirect('chartofaccountsoverview',pv)


def add_account_comment(request,pc):                                                                #new by tinto mt

    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
        company_id = request.session['login_id']
        if request.method=="POST":
                    
                    com=chart_of_accounts_comments()
                    c = CompanyDetails.objects.get(login_details=company_id)
            
                    comment_comments=request.POST['comment']
                    com.company=c
                    com.logindetails=log_user
                    com.comments=comment_comments
                    acc=Chart_of_Accounts.objects.get(id=pc)
                    com.chart_of_accounts=acc
                    
                    com.save()
                    return redirect('chartofaccountsoverview',pc)

    elif log_user.user_type == 'Staff':
        staff_id = request.session['login_id']
        if request.method=='POST':
            com=chart_of_accounts_comments()
            staff = LoginDetails.objects.get(id=staff_id)
            sf = StaffDetails.objects.get(login_details=staff)
            c=sf.company
            
            comment_comments=request.POST['comment']
            com.company=c
            com.logindetails=log_user
            com.comments=comment_comments
            acc=Chart_of_Accounts.objects.get(id=pc)
            com.chart_of_accounts=acc
                    
            com.save()
            return redirect('chartofaccountsoverview',pc)



from django.db.models import Max

def account_view_sort_by_name(request, pk):                                                                #new by tinto mt
    # Retrieve all items and convert them to a list of dictionaries
    acc = Chart_of_Accounts.objects.all().order_by('account_name')
    selacc = Chart_of_Accounts.objects.get(id=pk)
    est_comments = chart_of_accounts_comments.objects.filter(chart_of_accounts=pk)

    latest_date = Chart_of_Accounts_History.objects.filter(chart_of_accounts_id=pk).aggregate(latest_date=Max('Date'))['latest_date']
    filtered_data = Chart_of_Accounts_History.objects.filter(Date=latest_date, chart_of_accounts_id=pk)

    return render(request, 'zohomodules/chartofaccounts/chartofaccountsoverview.html', {'acc': acc, 'selacc': selacc, 'latest_item_id': filtered_data, 'est_comments': est_comments})


def shareaccountToEmail(request,pt):                                                                #new by tinto mt
    if request.user: 
        try:
            if request.method == 'POST':
                emails_string = request.POST['email_ids']
                # Split the string by commas and remove any leading or trailing whitespace
                emails_list = [email.strip() for email in emails_string.split(',')]
                email_message = request.POST['email_message']
                print(emails_list)
                print('1')
           
           
                acc = Chart_of_Accounts.objects.get(id=pt)
                context = {
                
                    'selacc':acc,
                }
                print('2')
                template_path = 'zohomodules/chartofaccounts/accountemailpdf.html'
                print('3')
                template = get_template(template_path)
                print('4')
                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)#, link_callback=fetch_resources)
                pdf = result.getvalue()
                print('5')
                filename = f'Account Details.pdf'
                subject = f"Account"
                email = EmailMessage(subject, f"Hi,\nPlease find the attached Account Details. \n{email_message}\n\n--\nRegards,\n{acc.account_name}\n{acc.account_type}", from_email=settings.EMAIL_HOST_USER,to=emails_list)
                email.attach(filename,pdf,"application/pdf")
                email.send(fail_silently=False)
                msg = messages.success(request, 'Details has been shared via email successfully..!')
                return redirect(chartofaccountsoverview,pt)
        except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect(chartofaccountsoverview,pt) 
        
        #--------------------------------------------------- TINTO VIEW CHART OF ACCOUNTS END-------------------------------------------








        



def filter_item_view_Active(request,pk): 
        if 'login_id' in request.session:
            login_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=login_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                items=Items.objects.filter(activation_tag='Active',company=dash_details.company)  
                selitem=Items.objects.get(id=pk)
                est_comments=Items_comments.objects.filter(Items=pk)
                stock_value=selitem.opening_stock*selitem.purchase_price  
                latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
                filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                context = {
                        'details': dash_details,
                        'items':items,
                        'selitem':selitem,
                        'stock_value':stock_value,
                        'latest_item_id':filtered_data,
                        'est_comments':est_comments,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/items/itemsoverview.html',context)
        if log_details.user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details)
                items=Items.objects.filter(activation_tag='Active',company=dash_details)  
                selitem=Items.objects.get(id=pk)
                est_comments=Items_comments.objects.filter(Items=pk)
                stock_value=selitem.opening_stock*selitem.purchase_price  
                latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
                filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)
      
                allmodules= ZohoModules.objects.get(company=dash_details,status='New')
                context = {
                        'details': dash_details,
                        'items':items,
                        'selitem':selitem,
                        'stock_value':stock_value,
                        'latest_item_id':filtered_data,
                        'est_comments':est_comments,
                        'allmodules': allmodules,
                }
        return render(request,'zohomodules/items/itemsoverview.html',context)

                                                                  #new by tinto mt

def filter_item_view_inActive(request,pk): 
        if 'login_id' in request.session:
            login_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=login_id)
        if log_details.user_type == 'Staff':
                items=Items.objects.filter(activation_tag='inactive',company=dash_details.company)  
                dash_details = StaffDetails.objects.get(login_details=log_details)
                selitem=Items.objects.get(id=pk)
                est_comments=Items_comments.objects.filter(Items=pk)
                stock_value=selitem.opening_stock*selitem.purchase_price  
                latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
                filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                context = {
                        'details': dash_details,
                        'items':items,
                        'selitem':selitem,
                        'stock_value':stock_value,
                        'latest_item_id':filtered_data,
                        'est_comments':est_comments,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/items/itemsoverview.html',context)
        if log_details.user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details)
                items=Items.objects.filter(activation_tag='inactive',company=dash_details)  
                selitem=Items.objects.get(id=pk)
                est_comments=Items_comments.objects.filter(Items=pk)
                stock_value=selitem.opening_stock*selitem.purchase_price  
                latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
                filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)
      
                allmodules= ZohoModules.objects.get(company=dash_details,status='New')
                context = {
                        'details': dash_details,
                        'items':items,
                        'selitem':selitem,
                        'stock_value':stock_value,
                        'latest_item_id':filtered_data,
                        'est_comments':est_comments,
                        'allmodules': allmodules,
                }
        return render(request,'zohomodules/items/itemsoverview.html',context)                                                               #new by tinto mt

def create_item(request):                                                                #new by tinto mt
    
    login_id = request.session['login_id']
    if 'login_id' not in request.session:
        return redirect('/')
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
        company_id = request.session['login_id']
        
        if request.method=='POST':
            a=Items()
            b=Item_Transaction_History()
            c = CompanyDetails.objects.get(login_details=company_id)
            b.company=c
            b.logindetails=log_user
            a.login_details=log_user
            a.company=c
            b.Date=date.today()
            a.item_type = request.POST.get("type",None)
            a.item_name = request.POST.get("name",None)
            unit_id = request.POST.get("unit")
            uid=Unit.objects.get(id=unit_id)
            # unit_instance = get_object_or_404(Unit, id=unit_id)
            a.unit = uid
            a.hsn_code = request.POST.get("hsn",None)
            a.tax_reference = request.POST.get("radio",None)
            a.intrastate_tax = request.POST.get("intra",None)
            a.interstate_tax= request.POST.get("inter",None)
            a.selling_price = request.POST.get("sel_price",None)
            a.sales_account = request.POST.get("sel_acc",None)
            a.sales_description = request.POST.get("sel_desc",None)
            a.purchase_price = request.POST.get("cost_price",None)
            a.purchase_account = request.POST.get("cost_acc",None)
            a.purchase_description = request.POST.get("pur_desc",None)
            a.minimum_stock_to_maintain = request.POST.get("minimum_stock",None)
            a.activation_tag = 'Active'
            a.type = 'Opening Stock'
            a.inventory_account = request.POST.get("invacc",None)
            a.opening_stock = request.POST.get("openstock",None)
            a.opening_stock_per_unit = request.POST.get("rate",None)
            
            a.save()    
            t=Items.objects.get(id=a.id)
            b.items=t
            b.save()
            return redirect('items_list')
    elif log_user.user_type == 'Staff':
        staff_id = request.session['login_id']
        if request.method=='POST':
            a=Items()
            b=Item_Transaction_History()
            staff = LoginDetails.objects.get(id=staff_id)
            sf = StaffDetails.objects.get(login_details=staff)
            c=sf.company
            b.Date=date.today()
            b.company=c
            b.logindetails=log_user
            a.login_details=log_user
            a.company=c
            a.item_type = request.POST.get("type",None)
            a.item_name = request.POST.get("name",None)
            unit_id = request.POST.get("unit")
            unit_instance = get_object_or_404(Unit, id=unit_id)
            a.unit = unit_instance
            a.hsn_code = request.POST.get("hsn",None)
            a.tax_reference = request.POST.get("radio",None)
            a.intrastate_tax = request.POST.get("intra",None)
            a.interstate_tax= request.POST.get("inter",None)
            a.selling_price = request.POST.get("sel_price",None)
            a.sales_account = request.POST.get("sel_acc",None)
            a.sales_description = request.POST.get("sel_desc",None)
            a.purchase_price = request.POST.get("cost_price",None)
            a.purchase_account = request.POST.get("cost_acc",None)
            a.purchase_description = request.POST.get("pur_desc",None)
            a.minimum_stock_to_maintain = request.POST.get("minimum_stock",None)
            a.activation_tag = request.POST.get("status",None)
            a.inventory_account = request.POST.get("invacc",None)
            a.opening_stock = request.POST.get("openstock",None)
        
        

        
            a.save()    
            t=Items.objects.get(id=a.id)
            b.items=t
            b.save()
            return redirect('items_list')
    return redirect('items_list')



def chartofaccounts(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            acc=Chart_of_Accounts.objects.filter(company=dash_details)
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)

    
def create_account(request):                                                                #new by tinto mt
    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
        company_id = request.session['login_id']
        if request.method=='POST':
            a=Chart_of_Accounts()
            b=Chart_of_Accounts_History()
            account=Chart_of_Accounts.objects.all()
            c = CompanyDetails.objects.get(login_details=company_id)
            b.company=c
            b.logindetails=log_user
            b.action="Created"
            b.Date=date.today()
            a.login_details=log_user
            a.company=c
          
        
            a.account_type = request.POST.get("account_type",None)
            a.account_name = request.POST.get("account_name",None)
            a.account_code = request.POST.get("account_code",None)
            a.account_number = request.POST.get("account_number2",None)
            a.account_description = request.POST.get("description",None)
            if a.account_type=="Other Current Assets":

                a.sub_account = request.POST.get("sub_account",None)
                a.parent_account = request.POST.get("parent_account",None)
               

            if a.account_type=="Cash":
             
                a.sub_account = request.POST.get("sub_account22",None)
                a.parent_account = request.POST.get("parent_account22",None)
               

            if a.account_type=="Fixed Assets":
            
                a.sub_account = request.POST.get("sub_account33",None)
                a.parent_account = request.POST.get("parent_account33",None)
               
            
            if a.account_type=="Stock":
               
                a.sub_account = request.POST.get("sub_account44",None)
                a.parent_account = request.POST.get("parent_account44",None)
             
            
            if a.account_type=="Other Current Liability":
             
                a.sub_account = request.POST.get("sub_account55",None)
                a.parent_account = request.POST.get("parent_account55",None)
               
            if a.account_type=="Long Term Liability":
            
                a.sub_account = request.POST.get("sub_account66",None)
                a.parent_account = request.POST.get("parent_account66",None)
              
            
            if a.account_type=="Other Liability":
              
                a.sub_account = request.POST.get("sub_account77",None)
                a.parent_account = request.POST.get("parent_account77",None)
              
            if a.account_type=="Equity":
            
                a.sub_account = request.POST.get("sub_account88",None)
                a.parent_account = request.POST.get("parent_account88",None)
            
            
            if a.account_type=="Income":
             
                a.sub_account = request.POST.get("sub_account99",None)
                a.parent_account = request.POST.get("parent_account99",None)
              
            
            if a.account_type=="Expense":
             
                a.sub_account = request.POST.get("sub_account100",None)
                a.parent_account = request.POST.get("parent_account100",None)
              
            if a.account_type=="Cost Of Goods Sold":
              
                a.sub_account = request.POST.get("sub_account111",None)
                a.parent_account = request.POST.get("parent_account111",None)
             
            if a.account_type=="Other Expense":
             
                a.sub_account = request.POST.get("sub_account222",None)
                a.parent_account = request.POST.get("parent_account222",None)
               
            account_type=request.POST.get("account_type",None)
            if account_type == 'Other Assets':
                a.description = 'Track special assets like goodwill and other intangible assets'
            if account_type == 'Other Current Assets':
                a.description = 'Any short term asset that can be converted into cash or cash equivalents easily Prepaid expenses Stocks and Mutual Funds'
            if account_type == 'Cash':
                a.description = 'To keep track of cash and other cash equivalents like petty cash, undeposited funds, etc., use an organized accounting system  financial software'
            if account_type == 'Bank':
                a.description = 'To keep track of bank accounts like Savings, Checking, and Money Market accounts.'
            if account_type == 'Fixed Asset':
                a.description = 'Any long-term investment or asset that cannot be easily converted into cash includes: Land and Buildings, Plant, Machinery, and Equipment, Computers, Furniture.'
            if account_type == 'Stock':
                a.description = 'To keep track of your inventory assets.'
            if account_type == 'Payment Clearing':
                a.description = 'To keep track of funds moving in and out via payment processors like Stripe, PayPal, etc.'
            if account_type == 'Other Liability':
                a.description = 'Obligation of an entity arising from past transactions or events which would require repayment.Tax to be paid Loan to be Repaid Accounts Payableetc.'
            if account_type == 'Other Current Liability':
                a.description = 'Any short term liability like: Customer Deposits Tax Payable'
            if account_type == 'Credit Card':
                a.description = 'Create a trail of all your credit card transactions by creating a credit card account.'
            if account_type == 'Long Term Liability':
                a.description = 'Liabilities that mature after a minimum period of one year like: Notes Payable Debentures Long Term Loans '
            if account_type == 'Overseas Tax Payable':
                a.description = 'Track your taxes in this account if your business sells digital services to foreign customers.'
            if account_type == 'Equity':
                a.description = 'Owners or stakeholders interest on the assets of the business after deducting all the liabilities.'
            if account_type == 'Income':
                a.description = 'Income or Revenue earned from normal business activities like sale of goods and services to customers.'
            if account_type == 'Other Income':
                a.description = 'Income or revenue earned from activities not directly related to your business like : Interest Earned Dividend Earned'
            if account_type == 'Expense':
                a.description = 'Reflects expenses incurred for running normal business operations, such as : Advertisements and Marketing Business Travel Expenses License Fees Utility Expenses'
            if account_type == 'Cost Of Goods Sold':
                a.description = 'This indicates the direct costs attributable to the production of the goods sold by a company such as: Material and Labor costs Cost of obtaining raw materials'
            if account_type == 'Other Expense':
                a.description = 'Track miscellaneous expenses incurred for activities other than primary business operations or create additional accounts to track default expenses like insurance or contribution towards charity.'
       

            
    
            a.Create_status="active"
            ac_name=request.POST.get("account_name",None)
            if Chart_of_Accounts.objects.filter(account_name=ac_name,company=c).exists():
                error='yes'
                messages.error(request,'Account with same name exsits !!!')
                return redirect('addchartofaccounts')
            else:
                a.save()
                t=Chart_of_Accounts.objects.get(id=a.id)
                b.chart_of_accounts=t
                b.save()
                return redirect('chartofaccounts')
    elif log_user.user_type == 'Staff':
        staff_id = request.session['login_id']
        if request.method=='POST':
            a=Chart_of_Accounts()
            b=Chart_of_Accounts_History()
            staff = LoginDetails.objects.get(id=staff_id)
            sf = StaffDetails.objects.get(login_details=staff)
            c=sf.company
            b.Date=date.today()
            b.company=c
            b.logindetails=log_user
            a.login_details=log_user
            a.company=c
          
        
            a.account_type = request.POST.get("account_type",None)
            a.account_name = request.POST.get("account_name",None)
            a.account_code = request.POST.get("account_code",None)
            a.description = request.POST.get("description",None)
            if a.account_type=="Other Current Assets":

                a.sub_account = request.POST.get("sub_account",None)
                a.parent_account = request.POST.get("parent_account",None)
               

            if a.account_type=="Cash":
             
                a.sub_account = request.POST.get("sub_account22",None)
                a.parent_account = request.POST.get("parent_account22",None)
               

            if a.account_type=="Fixed Assets":
            
                a.sub_account = request.POST.get("sub_account33",None)
                a.parent_account = request.POST.get("parent_account33",None)
               
            
            if a.account_type=="Stock":
               
                a.sub_account = request.POST.get("sub_account44",None)
                a.parent_account = request.POST.get("parent_account44",None)
             
            
            if a.account_type=="Other Current Liability":
             
                a.sub_account = request.POST.get("sub_account55",None)
                a.parent_account = request.POST.get("parent_account55",None)
               
            if a.account_type=="Long Term Liability":
            
                a.sub_account = request.POST.get("sub_account66",None)
                a.parent_account = request.POST.get("parent_account66",None)
              
            
            if a.account_type=="Other Liability":
              
                a.sub_account = request.POST.get("sub_account77",None)
                a.parent_account = request.POST.get("parent_account77",None)
              
            if a.account_type=="Equity":
            
                a.sub_account = request.POST.get("sub_account88",None)
                a.parent_account = request.POST.get("parent_account88",None)
            
            
            if a.account_type=="Income":
             
                a.sub_account = request.POST.get("sub_account99",None)
                a.parent_account = request.POST.get("parent_account99",None)
              
            
            if a.account_type=="Expense":
             
                a.sub_account = request.POST.get("sub_account100",None)
                a.parent_account = request.POST.get("parent_account100",None)
              
            if a.account_type=="Cost Of Goods Sold":
              
                a.sub_account = request.POST.get("sub_account111",None)
                a.parent_account = request.POST.get("parent_account111",None)
             
            if a.account_type=="Other Expense":
             
                a.sub_account = request.POST.get("sub_account222",None)
                a.parent_account = request.POST.get("parent_account222",None)
               
            account_type=request.POST.get("account_type",None)
            if account_type == 'Other Assets':
                a.description = 'Track special assets like goodwill and other intangible assets'
            if account_type == 'Other Current Assets':
                a.description = 'Any short term asset that can be converted into cash or cash equivalents easily Prepaid expenses Stocks and Mutual Funds'
            if account_type == 'Cash':
                a.description = 'To keep track of cash and other cash equivalents like petty cash, undeposited funds, etc., use an organized accounting system  financial software'
            if account_type == 'Bank':
                a.description = 'To keep track of bank accounts like Savings, Checking, and Money Market accounts.'
            if account_type == 'Fixed Asset':
                a.description = 'Any long-term investment or asset that cannot be easily converted into cash includes: Land and Buildings, Plant, Machinery, and Equipment, Computers, Furniture.'
            if account_type == 'Stock':
                a.description = 'To keep track of your inventory assets.'
            if account_type == 'Payment Clearing':
                a.description = 'To keep track of funds moving in and out via payment processors like Stripe, PayPal, etc.'
            if account_type == 'Other Liability':
                a.description = 'Obligation of an entity arising from past transactions or events which would require repayment.Tax to be paid Loan to be Repaid Accounts Payableetc.'
            if account_type == 'Other Current Liability':
                a.description = 'Any short term liability like: Customer Deposits Tax Payable'
            if account_type == 'Credit Card':
                a.description = 'Create a trail of all your credit card transactions by creating a credit card account.'
            if account_type == 'Long Term Liability':
                a.description = 'Liabilities that mature after a minimum period of one year like: Notes Payable Debentures Long Term Loans '
            if account_type == 'Overseas Tax Payable':
                a.description = 'Track your taxes in this account if your business sells digital services to foreign customers.'
            if account_type == 'Equity':
                a.description = 'Owners or stakeholders interest on the assets of the business after deducting all the liabilities.'
            if account_type == 'Income':
                a.description = 'Income or Revenue earned from normal business activities like sale of goods and services to customers.'
            if account_type == 'Other Income':
                a.description = 'Income or revenue earned from activities not directly related to your business like : Interest Earned Dividend Earned'
            if account_type == 'Expense':
                a.description = 'Reflects expenses incurred for running normal business operations, such as : Advertisements and Marketing Business Travel Expenses License Fees Utility Expenses'
            if account_type == 'Cost Of Goods Sold':
                a.description = 'This indicates the direct costs attributable to the production of the goods sold by a company such as: Material and Labor costs Cost of obtaining raw materials'
            if account_type == 'Other Expense':
                a.description = 'Track miscellaneous expenses incurred for activities other than primary business operations or create additional accounts to track default expenses like insurance or contribution towards charity.'
       
    
            a.Create_status="active"
            ac_name=request.POST.get("account_name",None)
            if Chart_of_Accounts.objects.filter(account_name=ac_name,company=c).exists():
                error='yes'
                messages.error(request,'Account with same name exsits')
                return redirect('addchartofaccounts')
            else:
                a.save()
                t=Chart_of_Accounts.objects.get(id=a.id)
                b.chart_of_accounts=t
                b.save()
                return redirect('chartofaccounts')

    return redirect('addchartofaccounts')

def chartofaccountsoverview(request,pk):                                                                #new by tinto mt
       if 'login_id' in request.session:
        login_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=login_id)
        if log_details.user_type == 'Staff':
                    dash_details = StaffDetails.objects.get(login_details=log_details)

                    allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                
                    acc=Chart_of_Accounts.objects.filter(company=dash_details.company)  
                    selacc=Chart_of_Accounts.objects.get(id=pk)  
                    est_comments=chart_of_accounts_comments.objects.filter(chart_of_accounts=pk)
                    latest_date = Chart_of_Accounts_History.objects.filter(chart_of_accounts_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
                    filtered_data = Chart_of_Accounts_History.objects.filter(Date=latest_date, chart_of_accounts_id=pk)
                    context = {
                        'details': dash_details,
                    
                        'allmodules': allmodules,
                        'acc':acc,
                        'selacc':selacc,
                        'latest_item_id':filtered_data,
                        'est_comments':est_comments,
                    }
                    return render(request, 'zohomodules/chartofaccounts/chartofaccountsoverview.html',context)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
       
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            acc=Chart_of_Accounts.objects.filter(company=dash_details)  
            selacc=Chart_of_Accounts.objects.get(id=pk)  
            est_comments=chart_of_accounts_comments.objects.filter(chart_of_accounts=pk)
            latest_date = Chart_of_Accounts_History.objects.filter(chart_of_accounts_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
            filtered_data = Chart_of_Accounts_History.objects.filter(Date=latest_date, chart_of_accounts_id=pk)
            context = {
                        'details': dash_details,
                    
                        'allmodules': allmodules,
                        'acc':acc,
                        'selacc':selacc,
                        'latest_item_id':filtered_data,
                        'est_comments':est_comments,
                    }
    
            return render(request, 'zohomodules/chartofaccounts/chartofaccountsoverview.html',context)






def create_item(request):                                                                #new by tinto mt
    
    login_id = request.session['login_id']
    if 'login_id' not in request.session:
        return redirect('/')
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
        company_id = request.session['login_id']
        
        if request.method=='POST':
            a=Items()
            b=Item_Transaction_History()
            c = CompanyDetails.objects.get(login_details=company_id)
            b.company=c
            b.logindetails=log_user
            a.login_details=log_user
            a.company=c
            a.item_type = request.POST.get("type",None)
            a.item_name = request.POST.get("name",None)
            unit_id = request.POST.get("unit")
            uid=Unit.objects.get(id=unit_id)
            # unit_instance = get_object_or_404(Unit, id=unit_id)
            a.unit = uid
            a.hsn_code = request.POST.get("hsn",None)
            a.tax_reference = request.POST.get("radio",None)
            a.intrastate_tax = request.POST.get("intra",None)
            a.interstate_tax= request.POST.get("inter",None)
            a.selling_price = request.POST.get("sel_price",None)
            a.sales_account = request.POST.get("sel_acc",None)
            a.sales_description = request.POST.get("sel_desc",None)
            a.purchase_price = request.POST.get("cost_price",None)
            a.purchase_account = request.POST.get("cost_acc",None)
            a.purchase_description = request.POST.get("pur_desc",None)
            minstock=request.POST.get("minimum_stock",None)
            if minstock != "":
                a.minimum_stock_to_maintain = request.POST.get("minimum_stock",None)
            else:
                a.minimum_stock_to_maintain = 0
            a.activation_tag = 'Active'
            a.type = 'Opening Stock'
            a.inventory_account = request.POST.get("invacc",None)
            a.opening_stock = request.POST.get("openstock",None)
            a.opening_stock_per_unit = request.POST.get("rate",None)
            
            a.save()    
            t=Items.objects.get(id=a.id)
            b.items=t
            b.save()
            return redirect('items_list')
    elif log_user.user_type == 'Staff':
        staff_id = request.session['login_id']
        if request.method=='POST':
            a=Items()
            b=Item_Transaction_History()
            staff = LoginDetails.objects.get(id=staff_id)
            sf = StaffDetails.objects.get(login_details=staff)
            c=sf.company
            b.Date=date.today()
            b.company=c
            b.logindetails=log_user
            a.login_details=log_user
            a.company=c
            a.item_type = request.POST.get("type",None)
            a.item_name = request.POST.get("name",None)
            unit_id = request.POST.get("unit")
            unit_instance = get_object_or_404(Unit, id=unit_id)
            a.unit = unit_instance
            a.hsn_code = request.POST.get("hsn",None)
            a.tax_reference = request.POST.get("radio",None)
            a.intrastate_tax = request.POST.get("intra",None)
            a.interstate_tax= request.POST.get("inter",None)
            a.selling_price = request.POST.get("sel_price",None)
            a.sales_account = request.POST.get("sel_acc",None)
            a.sales_description = request.POST.get("sel_desc",None)
            a.purchase_price = request.POST.get("cost_price",None)
            a.purchase_account = request.POST.get("cost_acc",None)
            a.purchase_description = request.POST.get("pur_desc",None)
            minstock=request.POST.get("minimum_stock",None)
            if minstock != "":
                a.minimum_stock_to_maintain = request.POST.get("minimum_stock",None)
            else:
                a.minimum_stock_to_maintain = 0
            # a.activation_tag = request.POST.get("status",None)
            a.inventory_account = request.POST.get("invacc",None)
            a.opening_stock = request.POST.get("openstock",None)
        
        

        
            a.save()    
            t=Items.objects.get(id=a.id)
            b.items=t
            b.save()
            return redirect('items_list')
    return redirect('items_list')






def delete_account_comment(request,ph,pr):                                                                #new by tinto mt
    acc=chart_of_accounts_comments.objects.filter(id=ph)
    acc.delete()
    ac=Chart_of_Accounts.objects.get(id=pr)
    
    return redirect(chartofaccountsoverview,ac.id)

def delete_item_comment(request,ph,pr):                                                                #new by tinto mt
    items=Items_comments.objects.filter(id=ph)
    items.delete()
    ac=Items.objects.get(id=pr)
    
    return redirect(itemsoverview,ac.id)

from django.db.models import Q
def accounts_asset_filter(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company,account_type__in=["Other Current Asset", "Fixed Asset","Other Asset"])
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            acc = Chart_of_Accounts.objects.filter(company=dash_details,account_type__in=["Other Current Asset", "Fixed Asset","Other Asset"])
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        

def accounts_liability_filter(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company,account_type__in=["Other Current Liability", "Other Liability","Long Term Liability","Credit card","Overseas Tax Payable"])
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            acc = Chart_of_Accounts.objects.filter(company=dash_details,account_type__in=["Other Current Liability", "Other Liability","Long Term Liability","Credit card","Overseas Tax Payable"])
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)


def accounts_equity_filter(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company,account_type__in=["Equity"])
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            acc = Chart_of_Accounts.objects.filter(company=dash_details,account_type__in=["Equity"])
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        

def accounts_income_filter(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company,account_type__in=["Income","Other Income"])
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            acc = Chart_of_Accounts.objects.filter(company=dash_details,account_type__in=["Income","Other Income"])
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
def accounts_expense_filter(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company,account_type__in=["Expense","Other Expense","Cost of Goods Sold"])
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            acc = Chart_of_Accounts.objects.filter(company=dash_details,account_type__in=["Expense","Other Expense","Cost of Goods Sold"])
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)


def edititems(request, pr):                                                                #new by tinto mt
    if 'login_id' in request.session:
        login_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    
    # Retrieve the chart of accounts entry
    item = get_object_or_404(Items, id=pr)
    

    # Check if 'company_id' is in the session

    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
      
     
        dash_details = CompanyDetails.objects.get(login_details=log_user)
        units = Unit.objects.filter(company=dash_details)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        item = get_object_or_404(Items, id=pr)
        accounts=Chart_of_Accounts.objects.filter(company=dash_details)
        units = Unit.objects.filter(company=dash_details)
        context = {
                    'item': item,
                    'units':units,
                    'details': dash_details,
                   'accounts': accounts,
                    'allmodules': allmodules,
            }
       
    
        
        if request.method=='POST':
   
            b=Item_Transaction_History()
            # c = CompanyDetails.objects.get(login_details=company_id)
            b.company=dash_details
            b.logindetails=log_user
            b.action="Edited"
            b.Date=date.today()
            item.login_details=log_user
            item.company=dash_details
            item.item_type = request.POST.get("type",None)
            item.item_name = request.POST.get("name",None)
            unit_id = request.POST.get("unit")
            unit_instance = get_object_or_404(Unit, id=unit_id)
            item.unit = unit_instance
            item.hsn_code = request.POST.get("hsn",None)
            item.tax_reference = request.POST.get("radio",None)
            if request.POST.get("radio",None) == 'taxable':

                item.intrastate_tax = request.POST.get("intra",None)
                item.interstate_tax= request.POST.get("inter",None)
            elif request.POST.get("radio",None) == 'None-Taxable':
                item.intrastate_tax = 0
                item.interstate_tax= 0
            item.selling_price = request.POST.get("sel_price",None)
            item.sales_account = request.POST.get("sel_acc",None)
            item.sales_description = request.POST.get("sel_desc",None)
            item.purchase_price = request.POST.get("cost_price",None)
            item.purchase_account = request.POST.get("cost_acc",None)
            item.purchase_description = request.POST.get("pur_desc",None)
            item.minimum_stock_to_maintain = request.POST.get("minimum_stock",None)
            item.activation_tag = request.POST.get("status",None)
            item.inventory_account = request.POST.get("invacc",None)
            item.opening_stock = request.POST.get("openstock",None)
            item.opening_stock_per_unit = request.POST.get("rate",None)
            
            
            # Save the changes
            item.save()
            t=Items.objects.get(id=item.id)
            b.items=t
            b.save()
            # Redirect to another page after successful update
            return redirect('itemsoverview', pr)
        return render(request, 'zohomodules/items/edititems.html',context)
    if log_user.user_type == 'Staff':
        dash_details = StaffDetails.objects.get(login_details=log_user)
                
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        item = get_object_or_404(Items, id=pr)
        units = Unit.objects.filter(company=dash_details.company)
        accounts=Chart_of_Accounts.objects.filter(company=dash_details.company)
        context = {
                    'item': item,
                    'units':units,
                    'details': dash_details,
                    'accounts': accounts,
                   
                    'allmodules': allmodules,
            }
 
        if request.method=='POST':
            a=Items()
            b=Item_Transaction_History()

            c=dash_details.company
            b.company=c
            b.logindetails=log_user
            b.action="Edited"
            b.Date=date.today()
            a.login_details=log_user
            a.company=c
            item.item_type = request.POST.get("type",None)
            item.item_name = request.POST.get("name",None)
            unit_id = request.POST.get("unit")
            unit_instance = get_object_or_404(Unit, id=unit_id)
            item.unit = unit_instance
            item.hsn_code = request.POST.get("hsn",None)
            item.tax_reference = request.POST.get("radio",None)
            item.intrastate_tax = request.POST.get("intra",None)
            item.interstate_tax= request.POST.get("inter",None)
            item.selling_price = request.POST.get("sel_price",None)
            item.sales_account = request.POST.get("sel_acc",None)
            item.sales_description = request.POST.get("sel_desc",None)
            item.purchase_price = request.POST.get("cost_price",None)
            item.purchase_account = request.POST.get("cost_acc",None)
            item.purchase_description = request.POST.get("pur_desc",None)
            item.minimum_stock_to_maintain = request.POST.get("minimum_stock",None)
            item.activation_tag = request.POST.get("status",None)
            item.inventory_account = request.POST.get("invacc",None)
            item.opening_stock = request.POST.get("openstock",None)
            item.save()
            t=Items.objects.get(id=item.id)
            b.items=t
            b.save()

            return redirect('itemsoverview', pr)
 
        return render(request, 'zohomodules/items/edititems.html', context)
   